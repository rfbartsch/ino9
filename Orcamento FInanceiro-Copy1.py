#!/usr/bin/env python
# coding: utf-8

# In[1]:


##CALCULO DE AuC
import os
import pandas as pd
import openpyxl


# Função lambda para extrair a data de vencimento
# Função para extrair a data da coluna 'Origem'
def extract_date(origem):
    if origem.startswith('Diversificacao_'):
        return origem.split('_', 2)[1]
    elif origem.startswith('Venc_RF_'):
        return origem.split('_', 4)[3]
    elif origem.startswith('Custodia_OffShore_MI_'):
        return origem.split('_', 4)[3].split('.', 1)[0]
    elif origem.startswith('Renda_Fixa_MTM_Ágio_'):
        return origem.split('_', 5)[4].split('.', 1)[0]
    elif origem.startswith('Diversificador_XPUS_Adeppar_Portfolio_Holding_'):
        return origem.split('_', 6)[5].split('.', 1)[0]
    elif origem.startswith('Diversificador_XPUS_Adeppar_Fixed_Income_'):
        return origem.split('_', 6)[5].split('.', 1)[0]
    else:
        return origem
    


# Função para manter o primeiro valor da coluna 'Data' no agrupamento
def first_value(x):
    return x.iloc[0]

PASTA_DOWNLOADS = r'C:\Users\dados.100486\Downloads'

#Financeiro
PASTA_DIVERSIFICACAO_INOVE_CONSOL = r'C:\Users\dados.100486\OneDrive - XP Investimentos\Tech Inove\Base Dados\Diversificacao Inove\Consolidado'
novo_nome3 = f'Diversificacao_Inove.xlsx'
path_to_file = os.path.join(PASTA_DIVERSIFICACAO_INOVE_CONSOL, f'Diversificacao_Inove.xlsx')
df = pd.read_excel(path_to_file)

df = df[['Origem','COD_ASSESSOR','NET Modulo','Produto','Sub Produto']]

# Criar a coluna "Linha AuC" com valores iniciais vazios
df['Linha AuC'] = ''

# Criar a coluna 'Linha AuC' utilizando a lógica condicional do seu exemplo
df['Linha AuC'] = df.apply(lambda row: 'AuC Previdência' if row['Produto'] == 'Previdência'
                            else 'AuC Fundos Imobiliários' if row['Produto'] == 'Fundo Imobiliário'
                            else 'AuC COE' if row['Produto'] == 'Alternativo' and row['Sub Produto'] == 'COE'
                            else 'AuC Cripto' if row['Produto'] == 'Alternativo' and row['Sub Produto'] == 'XC - BROKERAGE'
                            else 'AuC Fundos' if row['Produto'] == 'Fundos' and 'Diversificacao_' in row['Origem']
                            else 'AuC Renda Variável' if row['Produto'] == 'Renda Variável' and 'Diversificacao_' in row['Origem']
                            else 'AuC Saldo em Conta' if row['Produto'] == 'Somente Financeiro' and 'Diversificacao_' in row['Origem']
                            else 'AuC Renda Fixa' if row['Produto'] == 'Renda Fixa' and ('Diversificacao_' in row['Origem'] or 'Venc_RF' in row['Origem'] or 'Renda_Fixa_MTM' in row['Origem'])
                            else 'AuC Alternativos' if row['Produto'] == 'Fundos' and 'Venc_RF' in row['Origem']
                            else 'AuC MI - Saldo em Conta' if row['Produto'] == 'Somente Financeiro' and 'Custodia_OffShore_' in row['Origem']
                            else 'AuC MI - Renda Fixa' if row['Produto'] == 'Renda Fixa' and 'Custodia_OffShore_' in row['Origem']
                            else 'AuC MI - Renda Variável' if row['Produto'] == 'Renda Variável' and 'Custodia_OffShore_' in row['Origem']
                            else 'AuC MI - Fund' if row['Produto'] == 'Fund' and 'Custodia_OffShore_' in row['Origem']
                            else 'AuC XPUS - Saldo em Conta' if row['Produto'] == 'Somente Financeiro' and 'Diversificador_XPUS_' in row['Origem']
                            else 'AuC XPUS - Renda Fixa' if row['Produto'] == 'Renda Fixa' and 'Diversificador_XPUS_' in row['Origem']
                            else 'AuC XPUS - Renda Variável' if row['Produto'] == 'Renda Variável' and 'Diversificador_XPUS_' in row['Origem']
                            else 'AuC XPUS - Fundos' if row['Produto'] == 'Fundos' and 'Diversificador_XPUS_' in row['Origem']
                            else 'AuC XPUS - Alternativo' if row['Produto'] == 'Alternativo' and 'Diversificador_XPUS_' in row['Origem']
                            else '', axis=1)

        

df['Data'] = ''


    
    

# Criar a coluna 'Data' utilizando a função 'extract_date'
df['Data'] = df['Origem'].apply(extract_date)
print(df)

# Agrupar o DataFrame por 'COD_ASSESSOR' e 'Linha AuC' e calcular a soma da coluna 'NET_Modulo'
# Mantendo a coluna 'Data' correspondente ao primeiro valor de cada grupo
df = df.groupby(['COD_ASSESSOR', 'Linha AuC']).agg({'NET Modulo': 'sum', 'Data': first_value}).reset_index()


## separa o mes ano na data
# Converter a coluna 'Data' para o formato de data
df['Data'] = pd.to_datetime(df['Data'], format='%Y%m%d')

# Criar a coluna 'Mes_Ref' no formato mmyyyy
df['Mes_Ref'] = df['Data'].dt.strftime('%m%Y')
## Concatena o novo df com o df ja salvo (Consolidado AuC)
PASTA_FINANCEIRO = r'C:\Users\dados.100486\OneDrive - XP Investimentos\Tech Inove\Base Dados\Financeiro'
novo_nome3 = f'Consolidado AuC.xlsx'
path_to_file = os.path.join(PASTA_FINANCEIRO, novo_nome3)
df_consol = pd.read_excel(path_to_file)

df = pd.concat([df_consol, df], ignore_index=True)

## Exclui os duplicados (mes, ano, assessor, linha)
# Criar o ID concatenando as colunas relevantes

df['ID'] = df['COD_ASSESSOR'].astype(str) + '_' + df['Linha AuC'] + '_' + df['Mes_Ref'].astype(str)

df['Data'] = pd.to_datetime(df['Data'], format='%d/%m/%Y')
df['Data'] = df['Data'].dt.to_period('M').dt.to_timestamp()

# Ordenar o DataFrame pela coluna 'Data' em ordem decrescente
df.sort_values(by='Data', ascending=False, inplace=True)

# Remover os duplicados com base no ID, mantendo a primeira ocorrência (que terá a maior Data)
df.drop_duplicates(subset='ID', keep='first', inplace=True)
## Copia valor e cola pra dentro do arquivo orcamento

df.to_excel(path_to_file, index = False)

from openpyxl import load_workbook
import pandas as pd

# Definir o caminho para o arquivo Excel
file_path = r"C:\Users\dados.100486\OneDrive - XP Investimentos\Tech Inove\Base Dados\Financeiro\Orcamento.xlsx"

# Carregar a pasta de trabalho existente
book = load_workbook(file_path)

# Definir o escritor pandas com a pasta de trabalho aberta
writer = pd.ExcelWriter(file_path, engine='openpyxl')
writer.book = book

# Definir o DataFrame que você deseja escrever
df_to_copy = df.iloc[:10000, :6]

# Escrever o DataFrame na aba especificada
df_to_copy.to_excel(writer, sheet_name="AuC - Assessores", index=False, startrow=0, startcol=0)

# Salvar a pasta de trabalho
writer.save()


# In[ ]:





# In[ ]:





# In[ ]:





# In[2]:


#PJ1
import os
import pandas as pd

def first_value(x):
    return x.iloc[0]

PASTA_FINANCEIRO = r'C:\Users\dados.100486\OneDrive - XP Investimentos\Tech Inove\Base Dados\Financeiro'

# Caminho para a pasta com os arquivos
pasta = r'C:\Users\dados.100486\OneDrive - XP Investimentos\Tech Inove\Base Dados\Financeiro\Receita Atual'

# Listar os arquivos que começam com "Rel_Mensal"
arquivos = [f for f in os.listdir(pasta) if f.startswith('Rel_Mensal')]

# Concatenar os arquivos em um único DataFrame
df_concatenado = pd.DataFrame()
for arquivo in arquivos:
    caminho_arquivo = os.path.join(pasta, arquivo)
    df_temporario = pd.read_excel(caminho_arquivo,header = 2)
    df_receita_investimentos = pd.concat([df_concatenado, df_temporario], ignore_index=True)


#merge da coluna Produto com Produto/Categoria
# manter linha Receita e CFA

novo_nome = f'Orcamento.xlsx'
path_to_file = os.path.join(PASTA_FINANCEIRO, novo_nome)
df_class_receita = pd.read_excel(path_to_file, sheet_name ="Classificacao Receitas")


df_receita_investimentos = pd.merge(df_receita_investimentos, df_class_receita[['Produto','Linha Receita','CFA']], left_on='Produto/Categoria', right_on='Produto', how='left')

df_receita_investimentos = df_receita_investimentos.rename(columns={'Código Assessor':'Assessor',
                        'Comissão (R$) Escritório':'Valor'})


# Lista de colunas a serem mantidas no DataFrame
colunas_desejadas = ['Assessor', 'Linha Receita', 'Valor', 'Data','CFA']

# Filtrar o DataFrame para CFA igual a 'S' e manter apenas as colunas desejadas
#df_receita_investimentos = df_receita_investimentos[df_receita_investimentos['CFA'] == 'S'][colunas_desejadas]
df_receita_investimentos = df_receita_investimentos[colunas_desejadas]

df_receita_investimentos['Valor'] = df_receita_investimentos['Valor'].astype(float)
# Agrupar o DataFrame por 'COD_ASSESSOR' e 'Linha AuC' e calcular a soma da coluna 'NET_Modulo'
# Mantendo a coluna 'Data' correspondente ao primeiro valor de cada grupo
df_receita_investimentos = df_receita_investimentos.groupby(['Assessor', 'Linha Receita']).agg({'Valor': 'sum', 'Data': first_value, 'CFA':first_value}).reset_index()

# Remover a letra 'A' do início da coluna 'Assessor'
df_receita_investimentos['Assessor'] = df_receita_investimentos['Assessor'].str.lstrip('A')



df_receita_investimentos['Data'] = df_receita_investimentos['Data'].astype(str)
## Exclui os duplicados (mes, ano, assessor, linha)
# Criar o ID concatenando as colunas relevantes
df_receita_investimentos['ID'] = df_receita_investimentos['Assessor'].astype(str) + '_' + df_receita_investimentos['Linha Receita'] + '_' + df_receita_investimentos['Data']

df_receita_investimentos['Data'] = pd.to_datetime(df_receita_investimentos['Data'])

# Converter a coluna 'Data' para o formato de data
df_receita_investimentos['Data'] = pd.to_datetime(df_receita_investimentos['Data'], format='%d/%m/%Y')

# Definir o dia como 01 para todas as datas
df_receita_investimentos['Data'] = df_receita_investimentos['Data'].dt.to_period('M').dt.to_timestamp()

# Ordenar o DataFrame pela coluna 'Data' em ordem decrescente
df_receita_investimentos.sort_values(by='Data', ascending=False, inplace=True)

df_receita_investimentos['Assessor'] = df_receita_investimentos['Assessor'].astype(str)




# In[3]:


#XPUS
PASTA_ARQUIVOS_PADROES = r'C:\Users\dados.100486\OneDrive - XP Investimentos\Tech Inove\Base Dados\Arquivos Padrões'

arquivos = [f for f in os.listdir(pasta) if f.startswith('XP US_') and f.endswith('.csv')]

# Concatenar os arquivos em um único DataFrame
df_xpus = pd.DataFrame()
for arquivo in arquivos:
    caminho_arquivo = os.path.join(pasta, arquivo)
    df_temporario = pd.read_csv(caminho_arquivo, sep=';',header = 1)  # Especifica o separador como ponto e vírgula
    date_str = os.path.basename(arquivo)[-10:]
    date_str = date_str.split(".")[0]+"01"
    df_temporario['Data'] = date_str  # Extrai os últimos 6 dígitos do nome do arquivo usando o módulo os.path
    
    df_xpus = pd.concat([df_xpus, df_temporario], ignore_index=True)
    
df_xpus['Categoria'] = 'XPUS'

novo_ass = f'dAssessores.xls'
filepath = os.path.join(PASTA_ARQUIVOS_PADROES, novo_ass)
df_assessores = pd.read_excel(filepath)


df_xpus = pd.merge(df_xpus, df_assessores[['NOME_USUARIO','CODIGO_INOVE_USUARIO']], left_on='Rótulos de Linha', right_on='NOME_USUARIO', how='left')

    

novo_nome = f'Orcamento.xlsx'
path_to_file = os.path.join(PASTA_FINANCEIRO, novo_nome)
df_class_receita = pd.read_excel(path_to_file, sheet_name ="Classificacao Receitas")


df_xpus = pd.merge(df_xpus, df_class_receita[['Produto','Linha Receita','CFA']], left_on='Categoria', right_on='Produto', how='left')

df_xpus = df_xpus.rename(columns={'CODIGO_INOVE_USUARIO':'Assessor','Unnamed: 4':'Valor'})


# Lista de colunas a serem mantidas no DataFrame
colunas_desejadas = ['Assessor', 'Linha Receita', 'Valor', 'Data','CFA']

# Filtrar o DataFrame para CFA igual a 'S' e manter apenas as colunas desejadas
#df_xpus = df_xpus[df_xpus['CFA'] == 'S'][colunas_desejadas]
df_xpus = df_xpus[colunas_desejadas]

df_xpus['Valor'] = df_xpus['Valor'].str.replace(' R$    ', '').str.replace('R$', '').str.replace('-', '').str.replace('.', '').str.replace(',', '.').astype(float)

# Agrupar o DataFrame por 'COD_ASSESSOR' e 'Linha AuC' e calcular a soma da coluna 'NET_Modulo'
# Mantendo a coluna 'Data' correspondente ao primeiro valor de cada grupo
df_xpus = df_xpus.groupby(['Assessor', 'Linha Receita']).agg({'Valor': 'sum', 'Data': first_value, 'CFA':first_value}).reset_index()

df_xpus['Assessor'] = df_xpus['Assessor'].astype(str)

# Remover a letra 'A' do início da coluna 'Assessor'
df_xpus['Assessor'] = df_xpus['Assessor'].str.lstrip('A')

df_xpus['Data'] = df_xpus['Data'].astype(str)
## Exclui os duplicados (mes, ano, assessor, linha)
# Criar o ID concatenando as colunas relevantes
df_xpus['ID'] = df_xpus['Assessor'].astype(str) + '_' + df_xpus['Linha Receita'] + '_' + df_xpus['Data']

# Converter a coluna 'Data' para o formato de data
df_xpus['Data'] = pd.to_datetime(df_xpus['Data'], format='%Y/%m/%d')

# Definir o dia como 01 para todas as datas
df_xpus['Data'] = df_xpus['Data'].dt.to_period('M').dt.to_timestamp()

# Ordenar o DataFrame pela coluna 'Data' em ordem decrescente
df_xpus.sort_values(by='Data', ascending=False, inplace=True)


df_xpus['Assessor'] = df_xpus['Assessor'].astype(str)
# Aplicar a função split para separar o código do assessor no primeiro ponto e selecionar a parte antes do ponto (índice 0)
df_xpus['Assessor'] = df_xpus['Assessor'].apply(lambda x: x.split('.')[0])



# In[4]:


def processar_arquivos(pasta, valor_inicio_nome):
    arquivos = [f for f in os.listdir(pasta) if f.startswith(valor_inicio_nome) and f.endswith('.csv')]
    
    df_resultante = pd.DataFrame()
    for arquivo in arquivos:
        caminho_arquivo = os.path.join(pasta, arquivo)
        df_temporario = pd.read_csv(caminho_arquivo, sep=';')  # Especifica o separador como ponto e vírgula
        df_resultante = pd.concat([df_resultante, df_temporario], ignore_index=True)
    
    novo_nome = f'Orcamento.xlsx'
    path_to_file = os.path.join(PASTA_FINANCEIRO, novo_nome)
    df_class_receita = pd.read_excel(path_to_file, sheet_name ="Classificacao Receitas")
    
    df_resultante = pd.merge(df_resultante, df_class_receita[['Produto','Linha Receita','CFA']], left_on='Categoria', right_on='Produto', how='left')
    
    df_resultante = df_resultante.rename(columns={'Código Assessor':'Assessor','Comissão Escritório':'Valor'})
    
    colunas_desejadas = ['Assessor', 'Linha Receita', 'Valor', 'Data','CFA']
    
    df_resultante = df_resultante[colunas_desejadas]
    
    df_resultante['Valor'] = df_resultante['Valor'].str.replace(',', '.')
    df_resultante['Valor'] = df_resultante['Valor'].str.replace('.', '')
    df_resultante['Valor'] = df_resultante['Valor'].astype(float)/100
    
    df_resultante = df_resultante.groupby(['Assessor', 'Linha Receita']).agg({'Valor': 'sum', 'Data': 'first', 'CFA':'first'}).reset_index()
    
    df_resultante['Assessor'] = df_resultante['Assessor'].str.lstrip('A')
    
    df_resultante['Data'] = df_resultante['Data'].astype(str)
    df_resultante['ID'] = df_resultante['Assessor'].astype(str) + '_' + df_resultante['Linha Receita'] + '_' + df_resultante['Data']
    df_resultante['Data'] = pd.to_datetime(df_resultante['Data'], format='%d/%m/%Y')
    df_resultante['Data'] = df_resultante['Data'].dt.to_period('M').dt.to_timestamp()
    df_resultante.sort_values(by='Data', ascending=False, inplace=True)
    
    df_resultante['Assessor'] = df_resultante['Assessor'].astype(str)
    
    return df_resultante

df_cambio = processar_arquivos(pasta, valor_inicio_nome='CÂMBIO_')
df_mi = processar_arquivos(pasta, valor_inicio_nome='MERCADO INTERNACIONAL_')
df_saude = processar_arquivos(pasta, valor_inicio_nome='BTR')
df_cripto = processar_arquivos(pasta, valor_inicio_nome='XTAGE_')
df_corretagem_xpvp = processar_arquivos(pasta, valor_inicio_nome='CO-CORRETAGEM')
df_xpcs = processar_arquivos(pasta, valor_inicio_nome='XPCS_')
df_dm10 = processar_arquivos(pasta, valor_inicio_nome='DM10')
df_credito = processar_arquivos(pasta, valor_inicio_nome='CRÉDITO')


# In[5]:


print(df_dm10)


# In[6]:


#Consolidador


novo_nome3 = f'Consolidado Receita.xlsx'
path_to_file = os.path.join(PASTA_FINANCEIRO, novo_nome3)
df_consol_receita = pd.read_excel(path_to_file)

df_consol_receita = pd.concat([df_consol_receita, df_mi,df_receita_investimentos,df_cambio,df_cripto,df_xpus,df_corretagem_xpvp,df_xpcs,df_credito,df_dm10], ignore_index=True)

# Remover os duplicados com base no ID, mantendo a primeira ocorrência (que terá a maior Data)
df_consol_receita.drop_duplicates(subset='ID', keep='first', inplace=True)

df_consol_receita = df_consol_receita[colunas_desejadas]


df_consol_receita.to_excel(path_to_file, index = False)

print(df_consol_receita)


# In[7]:


from openpyxl import load_workbook
import pandas as pd

# Definir o caminho para o arquivo Excel
file_path = r"C:\Users\dados.100486\OneDrive - XP Investimentos\Tech Inove\Base Dados\Financeiro\Orcamento.xlsx"

# Carregar a pasta de trabalho existente
book = load_workbook(file_path)

# Definir o escritor pandas com a pasta de trabalho aberta
writer = pd.ExcelWriter(file_path, engine='openpyxl')
writer.book = book

# Definir o DataFrame que você deseja escrever
df_to_copy = df_consol_receita.iloc[:10000, :6]

# Escrever o DataFrame na aba especificada
df_to_copy.to_excel(writer, sheet_name="Consolidado Receitas", index=False, startrow=0, startcol=0)

# Salvar a pasta de trabalho
writer.save()

print(df_to_copy)


# In[ ]:





# In[ ]:





# In[8]:


print(df_credito)


# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:






# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:




